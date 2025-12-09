

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import sqlite3
import datetime
import sys

# --- LIBRERÍAS DE EXCEL Y PDF (Asegúrate de que estén instaladas: pip install fpdf2 openpyxl pandas) ---
import pandas as pd
from fpdf import FPDF, XPos, YPos
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill


# --- 2. CONFIGURACIÓN DE LA BASE DE DATOS ---

DB_FILE = 'contabilidad.db'

def conectar_db():
    """Establece la conexión a la base de datos SQLite."""
    try:
        conn = sqlite3.connect(DB_FILE)
        return conn
    except sqlite3.Error as e:
        messagebox.showerror("Error Crítico", f"Error al conectar con la base de datos: {e}")
        sys.exit(1)

def inicializar_db(conn):
    """Crea las tablas necesarias y maneja la migración."""
    cursor = conn.cursor()
    
    # Creación de tabla Cuentas
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS Cuentas (
            id INTEGER PRIMARY KEY,
            nombre TEXT NOT NULL UNIQUE,
            tipo TEXT NOT NULL -- ACTIVO, PASIVO, CAPITAL, INGRESO, GASTO
        )
    """)
    # Creación de tabla Transacciones
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS Transacciones (
            id INTEGER PRIMARY KEY,
            fecha TEXT NOT NULL,
            descripcion TEXT NOT NULL,
            cuenta_debito_id INTEGER NOT NULL,
            cuenta_credito_id INTEGER NOT NULL,
            monto REAL NOT NULL,
            referencia_documento TEXT, 
            referencia_banco TEXT, 
            FOREIGN KEY (cuenta_debito_id) REFERENCES Cuentas(id),
            FOREIGN KEY (cuenta_credito_id) REFERENCES Cuentas(id)
        )
    """)
    
    # --- RUTINA DE MIGRACIÓN (Añade columnas si no existen) ---
    def columna_existe(nombre_tabla, nombre_columna):
        cursor.execute(f"PRAGMA table_info({nombre_tabla})")
        columnas = [info[1] for info in cursor.fetchall()]
        return nombre_columna in columnas

    # Migración de referencia_documento
    if not columna_existe('Transacciones', 'referencia_documento'):
        try:
            cursor.execute("ALTER TABLE Transacciones ADD COLUMN referencia_documento TEXT")
        except sqlite3.OperationalError: pass
    
    # Migración de referencia_banco
    if not columna_existe('Transacciones', 'referencia_banco'):
        try:
            cursor.execute("ALTER TABLE Transacciones ADD COLUMN referencia_banco TEXT")
        except sqlite3.OperationalError: pass
            
    conn.commit()


# --- 3. LÓGICA DE LA APLICACIÓN (Backend) ---

class ContabilidadApp:
    def __init__(self):
        self.conn = conectar_db()
        inicializar_db(self.conn)
        self.cursor = self.conn.cursor()
        self.tipos_cuenta = ["ACTIVO", "PASIVO", "CAPITAL", "INGRESO", "GASTO"]
        self.tipos_saldo_acreedor = ["PASIVO", "CAPITAL", "INGRESO"]
        
        # Mapa de nombres de cuentas a IDs
        self.nombre_a_id = self._generar_mapa_cuentas() 

    def _generar_mapa_cuentas(self):
        """Genera el diccionario {nombre_cuenta: id_cuenta}"""
        mapa = {}
        cuentas = self.obtener_cuentas()
        for id, nombre, tipo in cuentas:
            mapa[nombre.upper()] = id
        return mapa
        
    def _ejecutar_consulta(self, query, params=()):
        """Maneja la ejecución de consultas con manejo de errores."""
        try:
            self.cursor.execute(query, params)
            self.conn.commit()
            return True, None
        except sqlite3.IntegrityError as e:
            return False, f"Error de integridad: Ya existe una cuenta con ese nombre o ID de cuenta inválido. ({e})"
        except sqlite3.Error as e:
            return False, f"Error de base de datos: {e}"

    # --- Gestión de Cuentas ---

    def crear_cuenta(self, nombre, tipo):
        """Agrega una nueva cuenta contable, retorna éxito y mensaje."""
        nombre = nombre.strip().upper()
        tipo = tipo.strip().upper()
        if not nombre:
            return False, "El nombre de la cuenta no puede estar vacío."
        if tipo not in self.tipos_cuenta:
            return False, f"Tipo de cuenta inválido. Debe ser uno de: {', '.join(self.tipos_cuenta)}"
            
        exito, error = self._ejecutar_consulta("INSERT INTO Cuentas (nombre, tipo) VALUES (?, ?)", (nombre, tipo))
        if exito:
            self.nombre_a_id[nombre] = self.cursor.lastrowid
            return True, f"Cuenta '{nombre}' ({tipo}) creada con éxito."
        else:
            return False, error

    def obtener_cuentas(self):
        """Devuelve todas las cuentas contables."""
        self.cursor.execute("SELECT id, nombre, tipo FROM Cuentas ORDER BY tipo, nombre")
        return self.cursor.fetchall()

    def obtener_nombre_cuenta(self, id):
        """Función auxiliar para obtener el nombre de una cuenta por ID."""
        self.cursor.execute("SELECT nombre FROM Cuentas WHERE id = ?", (id,))
        resultado = self.cursor.fetchone()
        return resultado[0] if resultado else None
        
    def eliminar_cuenta(self, cuenta_id):
        """Elimina una cuenta. SOLO si no tiene transacciones asociadas."""
        self.cursor.execute("""
            SELECT COUNT(*) FROM Transacciones 
            WHERE cuenta_debito_id = ? OR cuenta_credito_id = ?
        """, (cuenta_id, cuenta_id))
        count = self.cursor.fetchone()[0]
        
        if count > 0:
            return False, "Error: No se puede eliminar la cuenta porque tiene transacciones registradas."

        nombre_cuenta = self.obtener_nombre_cuenta(cuenta_id) 
        exito, error = self._ejecutar_consulta("DELETE FROM Cuentas WHERE id = ?", (cuenta_id,))
        if exito:
            if nombre_cuenta and nombre_cuenta.upper() in self.nombre_a_id:
                 del self.nombre_a_id[nombre_cuenta.upper()]
            return True, "Cuenta eliminada con éxito."
        else:
            return False, error
            
    # --- Gestión de Transacciones ---

    def registrar_transaccion(self, descripcion, debito_id, credito_id, monto, ref_doc, ref_banco):
        """Registra una transacción de doble partida, incluyendo las referencias externas."""
        if debito_id == credito_id:
            return False, "Error: Las cuentas de Débito y Crédito no pueden ser las mismas."
        if monto <= 0:
            return False, "Error: El monto de la transacción debe ser positivo."

        fecha = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        query = """
            INSERT INTO Transacciones (
                fecha, descripcion, cuenta_debito_id, cuenta_credito_id, monto, 
                referencia_documento, referencia_banco
            ) 
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """
        params = (
            fecha, descripcion.strip(), debito_id, credito_id, monto, 
            ref_doc.strip() or None, ref_banco.strip() or None
        )
        
        exito, error = self._ejecutar_consulta(query, params)
        return exito, error

    def obtener_transacciones_diario(self):
        """Devuelve todas las transacciones con los nombres de las cuentas y referencias."""
        query = """
            SELECT 
                T.id, T.fecha, T.descripcion, T.monto,
                CD.nombre AS nombre_debito, 
                CC.nombre AS nombre_credito,
                T.referencia_documento, T.referencia_banco
            FROM Transacciones T
            JOIN Cuentas CD ON T.cuenta_debito_id = CD.id
            JOIN Cuentas CC ON T.cuenta_credito_id = CC.id
            ORDER BY T.fecha DESC
        """
        self.cursor.execute(query)
        return self.cursor.fetchall()
        
    def eliminar_transaccion(self, transaccion_id):
        """Elimina una transacción del Libro Diario por su ID."""
        exito, error = self._ejecutar_consulta("DELETE FROM Transacciones WHERE id = ?", (transaccion_id,))
        if exito:
            return True, f"Transacción ID {transaccion_id} eliminada con éxito."
        else:
            return False, f"Error al eliminar la transacción: {error}"


    # --- Importar Transacciones desde Excel ---
    def importar_transacciones_excel(self, ruta_archivo):
        """Lee un archivo Excel y registra las transacciones."""
        try:
            df = pd.read_excel(ruta_archivo)
            
            # Limpiar y estandarizar nombres de columnas para facilitar el mapeo
            df.columns = [col.strip().replace(' ', '_').replace('.', '').upper() for col in df.columns]
            
            required_cols = ['DESCRIPCION', 'CUENTA_DEBITO', 'CUENTA_CREDITO', 'MONTO']
            if not all(col in df.columns for col in required_cols):
                return False, f"El archivo debe contener las columnas: {', '.join(required_cols)}"

            ref_doc_col = 'REF_DOC' if 'REF_DOC' in df.columns else None
            ref_banco_col = 'REF_BANCO' if 'REF_BANCO' in df.columns else None

            transacciones_importadas = 0
            errores = []

            for index, row in df.iterrows():
                try:
                    desc = str(row['DESCRIPCION']).strip()
                    # Asegurarse de manejar NaN/None en MONTO
                    if pd.isna(row['MONTO']): raise ValueError("Monto vacío")
                    monto = float(row['MONTO'])
                    
                    debito_nombre = str(row['CUENTA_DEBITO']).strip().upper()
                    credito_nombre = str(row['CUENTA_CREDITO']).strip().upper()
                    
                    debito_id = self.nombre_a_id.get(debito_nombre)
                    credito_id = self.nombre_a_id.get(credito_nombre)
                    
                    if not debito_id:
                        errores.append(f"Fila {index + 2}: Cuenta de Débito '{debito_nombre}' no existe.")
                        continue
                    if not credito_id:
                        errores.append(f"Fila {index + 2}: Cuenta de Crédito '{credito_nombre}' no existe.")
                        continue
                    
                    # Manejar referencias opcionales y NaN
                    ref_doc = str(row[ref_doc_col]).strip() if ref_doc_col and pd.notna(row[ref_doc_col]) else None
                    ref_banco = str(row[ref_banco_col]).strip() if ref_banco_col and pd.notna(row[ref_banco_col]) else None
                    
                    if monto <= 0:
                        errores.append(f"Fila {index + 2}: Monto debe ser positivo.")
                        continue
                    
                    # Registrar la transacción
                    fecha = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    query = """
                        INSERT INTO Transacciones (
                            fecha, descripcion, cuenta_debito_id, cuenta_credito_id, monto, 
                            referencia_documento, referencia_banco
                        ) 
                        VALUES (?, ?, ?, ?, ?, ?, ?)
                    """
                    params = (fecha, desc, debito_id, credito_id, monto, ref_doc, ref_banco)
                    self.cursor.execute(query, params)
                    transacciones_importadas += 1

                except ValueError:
                    errores.append(f"Fila {index + 2}: Error en el formato del Monto (debe ser numérico).")
                except Exception as e:
                    errores.append(f"Fila {index + 2}: Error desconocido: {e}")
                    
            self.conn.commit()
            
            if errores:
                mensaje_errores = "\n".join(errores[:10])
                if len(errores) > 10:
                    mensaje_errores += f"\n... y {len(errores) - 10} errores más."
                
                return False, f"Importación parcial. {transacciones_importadas} registradas con éxito, pero con {len(errores)} errores:\n{mensaje_errores}"
            
            return True, f"Importación exitosa. {transacciones_importadas} transacciones registradas."
            
        except FileNotFoundError:
            return False, "Error: Archivo no encontrado."
        except Exception as e:
            return False, f"Error al procesar el archivo Excel: {e}"

    # --- Reportes y Balances (Lógica de Cálculos) ---
    
    def calcular_saldos(self):
        """Calcula el saldo neto (Débito - Crédito) para cada cuenta."""
        self.cursor.execute("SELECT id, nombre, tipo FROM Cuentas")
        cuentas_data = self.cursor.fetchall()
        
        saldos = {}
        for id, nombre, tipo in cuentas_data:
            saldos[id] = {"nombre": nombre, "tipo": tipo, "saldo": 0.0}

        # Calcular Débitos totales por cuenta
        self.cursor.execute("SELECT cuenta_debito_id, SUM(monto) FROM Transacciones GROUP BY cuenta_debito_id")
        debitos = dict(self.cursor.fetchall())
        
        # Calcular Créditos totales por cuenta
        self.cursor.execute("SELECT cuenta_credito_id, SUM(monto) FROM Transacciones GROUP BY cuenta_credito_id")
        creditos = dict(self.cursor.fetchall())
        
        # Calcular saldo neto
        for id in saldos.keys():
            total_debito = debitos.get(id, 0.0)
            total_credito = creditos.get(id, 0.0)
            
            # Saldo = Débito - Crédito (Positivo=Deudor, Negativo=Acreedor)
            saldo_neto = total_debito - total_credito
            saldos[id]["saldo"] = saldo_neto
            
        return saldos
        
    def obtener_movimientos_cuenta(self, cuenta_id):
        """Devuelve todos los movimientos (Débito y Crédito) para una cuenta específica."""
        query = """
            SELECT 
                T.fecha, T.descripcion, T.cuenta_debito_id, T.cuenta_credito_id, T.monto,
                T.referencia_documento, T.referencia_banco
            FROM Transacciones T
            WHERE T.cuenta_debito_id = ? OR T.cuenta_credito_id = ?
            ORDER BY T.fecha ASC
        """
        self.cursor.execute(query, (cuenta_id, cuenta_id))
        return self.cursor.fetchall()

    def calcular_estado_resultados(self):
        """Calcula los resultados netos para el Estado de Resultados."""
        saldos = self.calcular_saldos()
        
        ingresos_totales = 0.0
        gastos_totales = 0.0
        
        for id, data in saldos.items():
            saldo = data["saldo"]
            tipo = data["tipo"]
            
            if tipo == "INGRESO":
                # Los ingresos tienen saldo acreedor (Crédito > Débito), por lo que el saldo neto es negativo.
                # Se toma el valor absoluto para sumar como valor positivo de Ingreso.
                ingresos_totales += abs(saldo)
            elif tipo == "GASTO":
                # Los gastos tienen saldo deudor (Débito > Crédito), por lo que el saldo neto es positivo.
                gastos_totales += saldo
        
        utilidad_perdida_neta = ingresos_totales - gastos_totales
        
        return {
            "ingresos_totales": ingresos_totales,
            "gastos_totales": gastos_totales,
            "utilidad_perdida_neta": utilidad_perdida_neta
        }

    def calcular_balance_general(self):
        """Calcula los totales para el Balance General (Ecuación Contable)."""
        saldos = self.calcular_saldos()
        
        total_activo = 0.0
        total_pasivo = 0.0
        total_capital = 0.0
        
        # Obtener la utilidad/pérdida del periodo
        estado_resultados = self.calcular_estado_resultados()
        utilidad_perdida_neta = estado_resultados['utilidad_perdida_neta']
        
        for id, data in saldos.items():
            saldo = data["saldo"]
            tipo = data["tipo"]
            
            if tipo == "ACTIVO":
                # Activo tiene saldo deudor (Débito > Crédito -> saldo > 0)
                total_activo += saldo
            elif tipo == "PASIVO":
                # Pasivo tiene saldo acreedor (Crédito > Débito -> saldo < 0)
                total_pasivo += abs(saldo)
            elif tipo == "CAPITAL":
                # Capital tiene saldo acreedor (Crédito > Débito -> saldo < 0)
                total_capital += abs(saldo)
        
        # El total de Patrimonio (Pasivo + Capital + Utilidad/Pérdida)
        total_patrimonio = total_pasivo + total_capital
        
        # Ajustar Capital con el Resultado del Ejercicio (Utilidad/Pérdida)
        # Una utilidad aumenta el patrimonio; una pérdida lo disminuye.
        total_capital_ajustado = total_capital + utilidad_perdida_neta
        total_pasivo_mas_capital_ajustado = total_pasivo + total_capital_ajustado

        return {
            "activo": total_activo,
            "pasivo": total_pasivo, # Pasivo sin resultado
            "capital": total_capital, # Capital sin resultado
            "utilidad_neta": utilidad_perdida_neta,
            "pasivo_mas_capital_ajustado": total_pasivo_mas_capital_ajustado # Activo = Pasivo + Capital + Utilidad
        }
    
    # --- Exportación a PDF (Completado) ---
    
    # 1. Exportación Balance de Comprobación a PDF (Ya estaba completo en tu envío)
    def exportar_balance_pdf(self, ruta_archivo):
        saldos = self.calcular_saldos()
        cuentas_con_saldo = [data for id, data in saldos.items() if abs(data["saldo"]) > 0.009]
        if not cuentas_con_saldo: return False, "No hay datos con saldo para exportar."
        
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", "B", 16)
        pdf.cell(0, 10, "Balance de Comprobación", new_x=XPos.LMARGIN, new_y=YPos.NEXT, align="C")
        pdf.set_font("Arial", "", 10)
        pdf.cell(0, 5, f"Fecha de Emisión: {datetime.date.today().strftime('%Y-%m-%d')}", new_x=XPos.LMARGIN, new_y=YPos.NEXT, align="C")
        pdf.ln(5)

        # Configuración de tabla
        pdf.set_fill_color(44, 62, 80) # Azul oscuro
        pdf.set_text_color(255, 255, 255) # Blanco
        pdf.set_font("Arial", "B", 9)
        
        col_widths = [60, 30, 40, 40] # Cuenta, Tipo, Débito, Crédito
        
        # Cabeceras
        headers = ["Cuenta", "Tipo", "Débito", "Crédito"]
        for i, header in enumerate(headers):
            pdf.cell(col_widths[i], 7, header, 1, 0, 'C', 1)
        pdf.ln()

        # Datos de la tabla
        pdf.set_text_color(0, 0, 0) # Negro
        pdf.set_font("Arial", "", 9)
        total_debitos_netos = 0.0
        total_creditos_netos = 0.0

        for data in cuentas_con_saldo:
            saldo = data["saldo"]
            debito_display = 0.00
            credito_display = 0.00
            
            if saldo > 0:
                debito_display = saldo
                total_debitos_netos += saldo
            elif saldo < 0:
                credito_display = abs(saldo)
                total_creditos_netos += credito_display
            
            pdf.cell(col_widths[0], 6, data["nombre"], 1, 0, 'L')
            pdf.cell(col_widths[1], 6, data["tipo"], 1, 0, 'C')
            pdf.cell(col_widths[2], 6, f"{debito_display:,.2f}", 1, 0, 'R')
            pdf.cell(col_widths[3], 6, f"{credito_display:,.2f}", 1, 0, 'R')
            pdf.ln()

        # Fila de Totales
        pdf.set_fill_color(189, 195, 199) # Gris claro
        pdf.set_text_color(0, 0, 0) # Negro
        pdf.set_font("Arial", "B", 9)
        pdf.cell(col_widths[0] + col_widths[1], 7, "TOTALES", 1, 0, 'R', 1) # Combine C1 y C2
        pdf.cell(col_widths[2], 7, f"{total_debitos_netos:,.2f}", 'BT', 0, 'R', 1)
        pdf.cell(col_widths[3], 7, f"{total_creditos_netos:,.2f}", 'BT', 0, 'R', 1)
        pdf.ln()

        try:
            pdf.output(ruta_archivo)
            return True, f"Balance de Comprobación exportado a PDF con éxito."
        except Exception as e:
            return False, f"Error al guardar el archivo PDF: {e}"

    # 2. Exportación Libro Diario a PDF (Ya estaba completo en tu envío)
    def exportar_diario_pdf(self, ruta_archivo):
        transacciones = self.obtener_transacciones_diario()
        if not transacciones: return False, "No hay transacciones para exportar en el Libro Diario."

        pdf = FPDF(orientation='L') # Horizontal
        pdf.add_page()
        pdf.set_font("Arial", "B", 16)
        pdf.cell(0, 10, "Libro Diario - Asientos Contables", new_x=XPos.LMARGIN, new_y=YPos.NEXT, align="C")
        pdf.set_font("Arial", "", 10)
        pdf.cell(0, 5, f"Fecha de Emisión: {datetime.date.today().strftime('%Y-%m-%d')}", new_x=XPos.LMARGIN, new_y=YPos.NEXT, align="C")
        pdf.ln(5)

        # Configuración de tabla
        pdf.set_fill_color(44, 62, 80) 
        pdf.set_text_color(255, 255, 255) 
        pdf.set_font("Arial", "B", 7)
        
        # Fecha | Ref. Doc | Ref. Banco | Descripción | Cuenta Débito | Cuenta Crédito | Monto
        col_widths = [18, 20, 20, 70, 45, 45, 25] 
        headers = ["Fecha", "Ref. Doc", "Ref. Banco", "Descripción", "Cuenta Débito", "Cuenta Crédito", "Monto"]
        
        for i, header in enumerate(headers):
            pdf.cell(col_widths[i], 5, header, 1, 0, 'C', 1)
        pdf.ln()

        # Datos de la tabla
        pdf.set_text_color(0, 0, 0)
        pdf.set_font("Arial", "", 7)
        
        for data in transacciones:
            id_trans, fecha, desc, monto, nombre_debito, nombre_credito, ref_doc, ref_banco = data
            
            fecha_display = fecha.split()[0]
            monto_str = f"{monto:,.2f}"
            
            # Fila DÉBITO
            pdf.cell(col_widths[0], 4, fecha_display, 1, 0, 'C')
            pdf.cell(col_widths[1], 4, ref_doc if ref_doc else '', 1, 0, 'L')
            pdf.cell(col_widths[2], 4, ref_banco if ref_banco else '', 1, 0, 'L')
            pdf.cell(col_widths[3], 4, desc, 1, 0, 'L')
            pdf.cell(col_widths[4], 4, nombre_debito, 1, 0, 'L')
            pdf.cell(col_widths[5], 4, '', 1, 0, 'L')
            pdf.cell(col_widths[6], 4, monto_str, 1, 0, 'R')
            pdf.ln()
            
            # Fila CRÉDITO (indentado visualmente en Descripción)
            pdf.cell(col_widths[0], 4, '', 1, 0, 'C')
            pdf.cell(col_widths[1], 4, '', 1, 0, 'L')
            pdf.cell(col_widths[2], 4, '', 1, 0, 'L')
            pdf.cell(col_widths[3], 4, '      a ' + nombre_credito, 1, 0, 'L')
            pdf.cell(col_widths[4], 4, '', 1, 0, 'L')
            pdf.cell(col_widths[5], 4, nombre_credito, 1, 0, 'L')
            pdf.cell(col_widths[6], 4, monto_str, 1, 0, 'R')
            pdf.ln()

        try:
            pdf.output(ruta_archivo)
            return True, f"Libro Diario exportado a PDF con éxito."
        except Exception as e:
            return False, f"Error al guardar el archivo PDF: {e}"
            
    # --- Exportación a Excel (Completo) ---
    
    def _configurar_celda_encabezado(self, cell):
        """Aplica estilo a las celdas de encabezado en Excel."""
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        borde = Side(style='thin', color="000000")
        cell.border = Border(top=borde, bottom=borde, left=borde, right=borde)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    def _configurar_celda_datos(self, cell, align='center', bold=False, border_bottom=False):
        """Aplica estilo a las celdas de datos en Excel."""
        cell.font = Font(bold=bold)
        cell.alignment = Alignment(horizontal=align, vertical="center")
        borde_lados = Side(style='thin', color="000000")
        borde_inferior = Side(style='double' if border_bottom else 'thin', color="000000")
        cell.border = Border(top=borde_lados, bottom=borde_inferior, left=borde_lados, right=borde_lados)

    def exportar_balance_excel(self, ruta_archivo):
        saldos = self.calcular_saldos()
        cuentas_con_saldo = [data for id, data in saldos.items() if abs(data["saldo"]) > 0.009]
        if not cuentas_con_saldo: return False, "No hay datos para exportar en el Balance."
        
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Balance Comprobacion"

            # Títulos y Cabeceras
            ws['A1'] = "Balance de Comprobación"
            ws['A1'].font = Font(size=14, bold=True)
            ws.merge_cells('A1:D1')
            ws['A1'].alignment = Alignment(horizontal="center")
            ws['A2'] = f"Fecha de Emisión: {datetime.date.today().strftime('%Y-%m-%d')}"
            ws.merge_cells('A2:D2')
            ws['A2'].alignment = Alignment(horizontal="center")
            
            headers = ["Cuenta", "Tipo", "Débito", "Crédito"]
            col_letras = ['A', 'B', 'C', 'D']
            
            for i, header in enumerate(headers):
                cell = ws[f'{col_letras[i]}4']
                cell.value = header
                self._configurar_celda_encabezado(cell)

            # Datos y Cálculos
            fila_inicio_datos = 5
            total_debitos_netos = 0.0
            total_creditos_netos = 0.0

            for i, data in enumerate(cuentas_con_saldo):
                fila = fila_inicio_datos + i
                saldo = data["saldo"]
                debito_display = 0.00
                credito_display = 0.00
                
                if saldo > 0:
                    debito_display = saldo
                    total_debitos_netos += saldo
                elif saldo < 0:
                    credito_display = abs(saldo)
                    total_creditos_netos += credito_display

                # Columna A: Cuenta
                cell_a = ws[f'A{fila}']
                cell_a.value = data["nombre"]
                self._configurar_celda_datos(cell_a, align='left')
                
                # Columna B: Tipo
                cell_b = ws[f'B{fila}']
                cell_b.value = data["tipo"]
                self._configurar_celda_datos(cell_b)

                # Columna C: Débito
                cell_c = ws[f'C{fila}']
                cell_c.value = debito_display
                cell_c.number_format = '#,##0.00'
                self._configurar_celda_datos(cell_c, align='right')

                # Columna D: Crédito
                cell_d = ws[f'D{fila}']
                cell_d.value = credito_display
                cell_d.number_format = '#,##0.00'
                self._configurar_celda_datos(cell_d, align='right')
                
            # Fila de Totales
            fila_total = fila_inicio_datos + len(cuentas_con_saldo)
            
            ws[f'A{fila_total}'].value = "TOTALES"
            ws.merge_cells(f'A{fila_total}:B{fila_total}')
            self._configurar_celda_datos(ws[f'A{fila_total}'], align='right', bold=True)
            self._configurar_celda_datos(ws[f'B{fila_total}'], align='right', bold=True) 

            cell_total_deb = ws[f'C{fila_total}']
            cell_total_deb.value = total_debitos_netos
            cell_total_deb.number_format = '#,##0.00'
            self._configurar_celda_datos(cell_total_deb, align='right', bold=True, border_bottom=True)
            
            cell_total_cred = ws[f'D{fila_total}']
            cell_total_cred.value = total_creditos_netos
            cell_total_cred.number_format = '#,##0.00'
            self._configurar_celda_datos(cell_total_cred, align='right', bold=True, border_bottom=True)
            
            # Ajustar anchos de columna
            ws.column_dimensions['A'].width = 30
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 18
            ws.column_dimensions['D'].width = 18

            wb.save(ruta_archivo)
            return True, "Balance de Comprobación exportado a Excel con éxito."
        except Exception as e:
            return False, f"Error al guardar el archivo Excel: {e}"

    def exportar_estado_resultados_excel(self, ruta_archivo):
        datos = self.calcular_estado_resultados()
        
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Estado Resultados"
            
            ws['A1'] = "Estado de Resultados"
            ws['A1'].font = Font(size=14, bold=True)
            ws.merge_cells('A1:B1')
            ws['A1'].alignment = Alignment(horizontal="center")
            ws['A2'] = f"Al {datetime.date.today().strftime('%Y-%m-%d')}"
            ws.merge_cells('A2:B2')
            ws['A2'].alignment = Alignment(horizontal="center")
            
            ws.column_dimensions['A'].width = 35
            ws.column_dimensions['B'].width = 20
            
            fila = 4
            
            def escribir_fila(concepto, monto, negrita=False, doble_borde=False):
                nonlocal fila
                cell_a = ws[f'A{fila}']
                cell_b = ws[f'B{fila}']
                
                cell_a.value = concepto
                self._configurar_celda_datos(cell_a, align='left', bold=negrita)
                
                cell_b.value = monto
                cell_b.number_format = '#,##0.00'
                self._configurar_celda_datos(cell_b, align='right', bold=negrita, border_bottom=doble_borde)
                fila += 1
                
            # Ingresos
            escribir_fila("Ingresos Totales:", datos["ingresos_totales"], negrita=True)
            
            # Menos Gastos
            ws[f'A{fila}'] = "(-) Gastos Totales:"
            ws[f'A{fila}'].font = Font(bold=True)
            ws[f'B{fila}'] = -datos["gastos_totales"] # Se pone negativo para resta visual
            ws[f'B{fila}'].number_format = '#,##0.00'
            borde = Side(style='thin', color="000000")
            ws[f'B{fila}'].border = Border(top=borde, bottom=borde, left=borde, right=borde)
            fila += 1
            
            # Utilidad
            escribir_fila("Utilidad/Pérdida Neta:", datos["utilidad_perdida_neta"], negrita=True, doble_borde=True)
            
            wb.save(ruta_archivo)
            return True, "Estado de Resultados exportado a Excel con éxito."
        except Exception as e:
            return False, f"Error al guardar el archivo Excel: {e}"

    def exportar_balance_general_excel(self, ruta_archivo):
        datos = self.calcular_balance_general()
        
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Balance General"
            
            ws['A1'] = "Balance General"
            ws['A1'].font = Font(size=14, bold=True)
            ws.merge_cells('A1:B1')
            ws['A1'].alignment = Alignment(horizontal="center")
            ws['A2'] = f"Al {datetime.date.today().strftime('%Y-%m-%d')}"
            ws.merge_cells('A2:B2')
            ws['A2'].alignment = Alignment(horizontal="center")
            
            ws.column_dimensions['A'].width = 35
            ws.column_dimensions['B'].width = 20
            
            fila = 4
            
            def escribir_fila(concepto, monto, negrita=False, doble_borde=False):
                nonlocal fila
                cell_a = ws[f'A{fila}']
                cell_b = ws[f'B{fila}']
                
                cell_a.value = concepto
                self._configurar_celda_datos(cell_a, align='left', bold=negrita)
                
                cell_b.value = monto
                cell_b.number_format = '#,##0.00'
                self._configurar_celda_datos(cell_b, align='right', bold=negrita, border_bottom=doble_borde)
                fila += 1
                
            # ACTIVO
            escribir_fila("ACTIVO TOTAL:", datos["activo"], negrita=True, doble_borde=True)
            fila += 1
            
            # PASIVO + CAPITAL
            escribir_fila("PASIVO TOTAL:", datos["pasivo"], negrita=True)
            escribir_fila("CAPITAL (Inicial):", datos["capital"], negrita=True)
            
            if datos["utilidad_neta"] > 0:
                 escribir_fila("(+) Utilidad del Ejercicio:", datos["utilidad_neta"])
            elif datos["utilidad_neta"] < 0:
                 escribir_fila("(-) Pérdida del Ejercicio:", datos["utilidad_neta"])
                 
            escribir_fila("PASIVO + CAPITAL TOTAL AJUSTADO:", datos["pasivo_mas_capital_ajustado"], negrita=True, doble_borde=True)

            wb.save(ruta_archivo)
            return True, "Balance General exportado a Excel con éxito."
        except Exception as e:
            return False, f"Error al guardar el archivo Excel: {e}"

    def cerrar_conexion(self):
        """Cierra la conexión con la base de datos."""
        self.conn.close()

# --- 4. INTERFAZ DE USUARIO (GUI con Tkinter) ---

class ContabilidadGUI:
    def __init__(self, root, app_backend):
        self.root = root
        self.app = app_backend
        self.root.title("Sistema Contable Integral para PYMES")
        self.root.geometry("1300x750")

        style = ttk.Style()
        style.theme_use('clam')
        style.configure('Treeview.Heading', font=('Inter', 10, 'bold'), background='#2c3e50', foreground='white')
        style.configure('Reporte.TFrame', background='#ecf0f1')
        style.configure('TButton', font=('Inter', 9, 'bold'), foreground='#2c3e50')
        style.map('TButton', background=[('active','#bdc3c7')])

        # Crear Pestañas (Notebook)
        self.notebook = ttk.Notebook(root)
        self.notebook.pack(pady=10, padx=10, expand=True, fill="both")

        # Frames para las Pestañas
        self.frame_transaccion = ttk.Frame(self.notebook, padding="10")
        self.frame_cuentas = ttk.Frame(self.notebook, padding="10")
        self.frame_diario = ttk.Frame(self.notebook, padding="10")
        self.frame_mayor = ttk.Frame(self.notebook, padding="10")
        self.frame_balance_comp = ttk.Frame(self.notebook, padding="10")
        self.frame_estado_resultados = ttk.Frame(self.notebook, padding="10")
        self.frame_balance_general = ttk.Frame(self.notebook, padding="10")
        self.frame_importar = ttk.Frame(self.notebook, padding="10") 

        # Agregar Pestañas en el orden deseado
        self.notebook.add(self.frame_transaccion, text="✍️ Registrar Transacción")
        self.notebook.add(self.frame_cuentas, text="📦 Gestión de Cuentas")
        self.notebook.add(self.frame_diario, text="📜 Libro Diario")
        self.notebook.add(self.frame_mayor, text="📖 Libro Mayor")
        self.notebook.add(self.frame_balance_comp, text="🧮 Balance de Comprobación")
        self.notebook.add(self.frame_estado_resultados, text="📈 Estado de Resultados")
        self.notebook.add(self.frame_balance_general, text="🏦 Balance General")
        self.notebook.add(self.frame_importar, text="📥 Importar Datos") 
        
        # Inicializar Pestañas
        self._setup_transaccion_tab()
        self._setup_cuentas_tab()
        self._setup_diario_tab()
        self._setup_mayor_tab()
        self._setup_balance_comp_tab()
        self._setup_estado_resultados_tab()
        self._setup_balance_general_tab()
        self._setup_importar_tab() 

        # Cargar datos iniciales
        self.actualizar_combos_transaccion()
        self.actualizar_lista_cuentas()
        self.actualizar_balance_comp()
        self.actualizar_libro_diario()
        self.actualizar_mayor_tab()
        self.actualizar_estado_resultados()
        self.actualizar_balance_general()
        
        # Configurar actualización automática al cambiar de pestaña
        self.notebook.bind("<<NotebookTabChanged>>", self._on_tab_change)

    # --- Actualización Global ---
    def recargar_vistas_contables(self):
        """Actualiza todas las vistas después de una creación/eliminación de transacción o cuenta."""
        self.actualizar_libro_diario()
        self.actualizar_mayor_tab()
        self.actualizar_balance_comp()
        self.actualizar_estado_resultados()
        self.actualizar_balance_general()

    def _on_tab_change(self, event):
        """Maneja la actualización de datos al cambiar de pestaña."""
        selected_tab = self.notebook.tab(self.notebook.select(), "text").strip()
        
        if "Balance de Comprobación" in selected_tab:
            self.actualizar_balance_comp()
        elif "Transacción" in selected_tab:
            self.actualizar_combos_transaccion()
        elif "Cuentas" in selected_tab:
            self.actualizar_lista_cuentas()
        elif "Libro Diario" in selected_tab:
            self.actualizar_libro_diario()
        elif "Libro Mayor" in selected_tab:
            self.actualizar_mayor_tab()
        elif "Estado de Resultados" in selected_tab:
            self.actualizar_estado_resultados()
        elif "Balance General" in selected_tab:
            self.actualizar_balance_general()
            
    # --- PESTAÑA: Registrar Transacción (Ya estaba) ---
    def _setup_transaccion_tab(self):
        frame = ttk.Frame(self.frame_transaccion)
        frame.pack(pady=20, padx=20)
        
        ttk.Label(frame, text="Nuevo Asiento Contable", font=('Inter', 16, 'bold')).grid(row=0, column=0, columnspan=2, pady=10)

        # Variables de control
        self.desc_var = tk.StringVar()
        self.monto_var = tk.DoubleVar(value=0.0)
        self.debito_var = tk.StringVar()
        self.credito_var = tk.StringVar()
        self.ref_doc_var = tk.StringVar()
        self.ref_banco_var = tk.StringVar()
        self.cuenta_map = {} # {Nombre: ID}

        # 1. Descripción
        ttk.Label(frame, text="Descripción/Concepto:", anchor='w').grid(row=1, column=0, sticky='w', padx=5, pady=5)
        ttk.Entry(frame, textvariable=self.desc_var, width=50).grid(row=1, column=1, sticky='ew', padx=5, pady=5)

        # 2. Monto
        ttk.Label(frame, text="Monto ($):", anchor='w').grid(row=2, column=0, sticky='w', padx=5, pady=5)
        ttk.Entry(frame, textvariable=self.monto_var, width=50).grid(row=2, column=1, sticky='ew', padx=5, pady=5)
        
        # 3. Cuenta Débito (Combobox)
        ttk.Label(frame, text="Cuenta Débito (Aumenta o Gasto):", anchor='w').grid(row=3, column=0, sticky='w', padx=5, pady=5)
        self.debito_combo = ttk.Combobox(frame, textvariable=self.debito_var, width=47, state='readonly')
        self.debito_combo.grid(row=3, column=1, sticky='ew', padx=5, pady=5)

        # 4. Cuenta Crédito (Combobox)
        ttk.Label(frame, text="Cuenta Crédito (Disminuye o Ingreso):", anchor='w').grid(row=4, column=0, sticky='w', padx=5, pady=5)
        self.credito_combo = ttk.Combobox(frame, textvariable=self.credito_var, width=47, state='readonly')
        self.credito_combo.grid(row=4, column=1, sticky='ew', padx=5, pady=5)
        
        # 5. Referencia Documento
        ttk.Label(frame, text="Ref. Documento (Factura, etc.):", anchor='w').grid(row=5, column=0, sticky='w', padx=5, pady=5)
        ttk.Entry(frame, textvariable=self.ref_doc_var, width=50).grid(row=5, column=1, sticky='ew', padx=5, pady=5)
        
        # 6. Referencia Banco
        ttk.Label(frame, text="Ref. Banco (Cheque, Transf., etc.):", anchor='w').grid(row=6, column=0, sticky='w', padx=5, pady=5)
        ttk.Entry(frame, textvariable=self.ref_banco_var, width=50).grid(row=6, column=1, sticky='ew', padx=5, pady=5)

        # Botón de registro
        ttk.Button(frame, text="Registrar Asiento (Partida Doble)", command=self.registrar_transaccion).grid(row=7, column=0, columnspan=2, pady=20)
        
    def actualizar_combos_transaccion(self):
        cuentas = self.app.obtener_cuentas()
        nombres_cuentas = []
        self.cuenta_map = {}
        
        for id, nombre, tipo in cuentas:
            nombre_display = f"{nombre} ({tipo})"
            nombres_cuentas.append(nombre_display)
            self.cuenta_map[nombre_display] = id

        self.debito_combo['values'] = nombres_cuentas
        self.credito_combo['values'] = nombres_cuentas
        
        # Limpiar selección si las cuentas ya no existen
        if self.debito_var.get() not in nombres_cuentas: self.debito_var.set("")
        if self.credito_var.get() not in nombres_cuentas: self.credito_var.set("")

    def registrar_transaccion(self):
        try:
            descripcion = self.desc_var.get()
            monto = self.monto_var.get()
            debito_nombre = self.debito_var.get()
            credito_nombre = self.credito_var.get()
            ref_doc = self.ref_doc_var.get()
            ref_banco = self.ref_banco_var.get()

            debito_id = self.cuenta_map.get(debito_nombre)
            credito_id = self.cuenta_map.get(credito_nombre)

            if not descripcion or not debito_id or not credito_id or monto <= 0:
                messagebox.showwarning("Datos Incompletos", "Debe completar la descripción, el monto (positivo) y seleccionar ambas cuentas.")
                return

            exito, mensaje = self.app.registrar_transaccion(descripcion, debito_id, credito_id, monto, ref_doc, ref_banco)

            if exito:
                messagebox.showinfo("Éxito", mensaje)
                # Limpiar campos
                self.desc_var.set("")
                self.monto_var.set(0.0)
                self.ref_doc_var.set("")
                self.ref_banco_var.set("")
                
                # Actualizar todas las vistas contables
                self.recargar_vistas_contables()
            else:
                messagebox.showerror("Error de Registro", mensaje)

        except ValueError:
            messagebox.showerror("Error de Monto", "El monto debe ser un número válido.")
        except Exception as e:
            messagebox.showerror("Error General", f"Ocurrió un error: {e}")

    # --- PESTAÑA: Gestión de Cuentas (Ya estaba) ---
    def _setup_cuentas_tab(self):
        # Frame principal
        frame = ttk.Frame(self.frame_cuentas)
        frame.pack(fill='both', expand=True)

        # Frame de registro de cuentas (arriba, 1/3 del espacio)
        reg_frame = ttk.LabelFrame(frame, text="✅ Registrar Nueva Cuenta", padding=10)
        reg_frame.pack(fill='x', padx=10, pady=5)
        
        # Variables de control
        self.nueva_cuenta_nombre_var = tk.StringVar()
        self.nueva_cuenta_tipo_var = tk.StringVar()

        # Inputs
        ttk.Label(reg_frame, text="Nombre:").grid(row=0, column=0, padx=5, pady=5, sticky='w')
        ttk.Entry(reg_frame, textvariable=self.nueva_cuenta_nombre_var, width=40).grid(row=0, column=1, padx=5, pady=5, sticky='ew')

        ttk.Label(reg_frame, text="Tipo:").grid(row=1, column=0, padx=5, pady=5, sticky='w')
        self.tipo_combo = ttk.Combobox(reg_frame, textvariable=self.nueva_cuenta_tipo_var, values=self.app.tipos_cuenta, state='readonly', width=37)
        self.tipo_combo.grid(row=1, column=1, padx=5, pady=5, sticky='ew')

        ttk.Button(reg_frame, text="Crear Cuenta", command=self.crear_cuenta).grid(row=0, column=2, rowspan=2, padx=15, pady=5, sticky='ns')

        reg_frame.grid_columnconfigure(1, weight=1)
        
        # Frame de listado (abajo, 2/3 del espacio)
        list_frame = ttk.Frame(frame)
        list_frame.pack(fill='both', expand=True, padx=10, pady=5)

        ttk.Label(list_frame, text="Cuentas Existentes", font=('Inter', 12, 'bold')).pack(pady=5)
        
        # Treeview para mostrar cuentas
        self.cuentas_tree = ttk.Treeview(list_frame, columns=("ID", "Nombre", "Tipo"), show="headings")
        self.cuentas_tree.heading("ID", text="ID", anchor=tk.CENTER)
        self.cuentas_tree.heading("Nombre", text="Nombre", anchor=tk.W)
        self.cuentas_tree.heading("Tipo", text="Tipo", anchor=tk.CENTER)
        
        self.cuentas_tree.column("ID", width=50, anchor=tk.CENTER)
        self.cuentas_tree.column("Nombre", width=400, anchor=tk.W)
        self.cuentas_tree.column("Tipo", width=100, anchor=tk.CENTER)
        
        self.cuentas_tree.pack(side='left', fill='both', expand=True)
        
        # Scrollbar
        scrollbar = ttk.Scrollbar(list_frame, orient="vertical", command=self.cuentas_tree.yview)
        self.cuentas_tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side='right', fill='y')

        # Botón de eliminación
        del_button_frame = ttk.Frame(list_frame)
        del_button_frame.pack(side='bottom', fill='x', pady=5)
        ttk.Button(del_button_frame, text="🗑️ Eliminar Cuenta Seleccionada", command=self.eliminar_cuenta).pack(side='left', padx=5)

    def crear_cuenta(self):
        nombre = self.nueva_cuenta_nombre_var.get()
        tipo = self.nueva_cuenta_tipo_var.get()
        
        exito, mensaje = self.app.crear_cuenta(nombre, tipo)
        
        if exito:
            messagebox.showinfo("Éxito", mensaje)
            self.nueva_cuenta_nombre_var.set("")
            self.nueva_cuenta_tipo_var.set("")
            self.actualizar_lista_cuentas()
            self.actualizar_combos_transaccion() 
        else:
            messagebox.showerror("Error", mensaje)

    def eliminar_cuenta(self):
        seleccion = self.cuentas_tree.focus()
        if not seleccion:
            messagebox.showwarning("Advertencia", "Debe seleccionar una cuenta para eliminar.")
            return

        cuenta_id = self.cuentas_tree.item(seleccion, 'values')[0]
        nombre_cuenta = self.cuentas_tree.item(seleccion, 'values')[1]

        if not messagebox.askyesno("Confirmar Eliminación", 
                                   f"¿Está seguro de eliminar la cuenta '{nombre_cuenta}'?\n\n¡Advertencia! Solo se podrá eliminar si no tiene transacciones registradas."):
            return

        exito, mensaje = self.app.eliminar_cuenta(int(cuenta_id))

        if exito:
            messagebox.showinfo("Éxito", mensaje)
            self.actualizar_lista_cuentas()
            self.actualizar_combos_transaccion()
            self.recargar_vistas_contables() # Importante: Recargar reportes después de eliminar cuenta
        else:
            messagebox.showerror("Error de Eliminación", mensaje)

    def actualizar_lista_cuentas(self):
        for item in self.cuentas_tree.get_children():
            self.cuentas_tree.delete(item)
            
        cuentas = self.app.obtener_cuentas()
        for id, nombre, tipo in cuentas:
            self.cuentas_tree.insert("", "end", values=(id, nombre, tipo))


    # --- PESTAÑA: Libro Diario (Completado) ---
    def _setup_diario_tab(self):
        frame = ttk.Frame(self.frame_diario)
        frame.pack(fill='both', expand=True, padx=10, pady=10)
        
        # Botones de Acción
        btn_frame = ttk.Frame(frame)
        btn_frame.pack(fill='x', pady=5)
        ttk.Button(btn_frame, text="🗑️ Eliminar Transacción Seleccionada", command=self.eliminar_transaccion).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="📄 Exportar a PDF", command=self.exportar_diario_pdf).pack(side='right', padx=5)
        
        # Treeview
        columns = ("ID", "Fecha", "Ref_Doc", "Ref_Banco", "Descripción", "Cuenta_Debito", "Cuenta_Credito", "Monto")
        self.diario_tree = ttk.Treeview(frame, columns=columns, show="headings")
        
        # Configuración de columnas
        self.diario_tree.heading("ID", text="ID")
        self.diario_tree.heading("Fecha", text="Fecha")
        self.diario_tree.heading("Ref_Doc", text="Ref. Doc.")
        self.diario_tree.heading("Ref_Banco", text="Ref. Banco")
        self.diario_tree.heading("Descripción", text="Descripción")
        self.diario_tree.heading("Cuenta_Debito", text="Cuenta Débito")
        self.diario_tree.heading("Cuenta_Credito", text="Cuenta Crédito")
        self.diario_tree.heading("Monto", text="Monto ($)")
        
        self.diario_tree.column("ID", width=40, anchor='center')
        self.diario_tree.column("Fecha", width=120, anchor='center')
        self.diario_tree.column("Ref_Doc", width=80, anchor='w')
        self.diario_tree.column("Ref_Banco", width=80, anchor='w')
        self.diario_tree.column("Descripción", width=250, anchor='w')
        self.diario_tree.column("Cuenta_Debito", width=150, anchor='w')
        self.diario_tree.column("Cuenta_Credito", width=150, anchor='w')
        self.diario_tree.column("Monto", width=100, anchor='e')
        
        self.diario_tree.pack(side='left', fill='both', expand=True)
        
        # Scrollbar
        scrollbar = ttk.Scrollbar(frame, orient="vertical", command=self.diario_tree.yview)
        self.diario_tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side='right', fill='y')

    def actualizar_libro_diario(self):
        for item in self.diario_tree.get_children():
            self.diario_tree.delete(item)
            
        transacciones = self.app.obtener_transacciones_diario()
        
        for trans in transacciones:
            id, fecha, desc, monto, nombre_debito, nombre_credito, ref_doc, ref_banco = trans
            # Formatear el monto con comas y dos decimales
            monto_str = f"{monto:,.2f}"
            
            self.diario_tree.insert("", "end", values=(
                id, fecha.split()[0], ref_doc if ref_doc else '', ref_banco if ref_banco else '', 
                desc, nombre_debito, nombre_credito, monto_str
            ))

    def eliminar_transaccion(self):
        seleccion = self.diario_tree.focus()
        if not seleccion:
            messagebox.showwarning("Advertencia", "Debe seleccionar una transacción para eliminar.")
            return

        transaccion_id = self.diario_tree.item(seleccion, 'values')[0]

        if not messagebox.askyesno("Confirmar Eliminación", 
                                   f"¿Está seguro de eliminar la Transacción ID {transaccion_id}?"):
            return

        exito, mensaje = self.app.eliminar_transaccion(int(transaccion_id))

        if exito:
            messagebox.showinfo("Éxito", mensaje)
            self.recargar_vistas_contables()
        else:
            messagebox.showerror("Error de Eliminación", mensaje)

    def exportar_diario_pdf(self):
        ruta_archivo = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("Archivos PDF", "*.pdf")],
            title="Guardar Libro Diario en PDF"
        )
        if ruta_archivo:
            exito, mensaje = self.app.exportar_diario_pdf(ruta_archivo)
            if exito:
                messagebox.showinfo("Éxito de Exportación", mensaje)
            else:
                messagebox.showerror("Error de Exportación", mensaje)


    # --- PESTAÑA: Libro Mayor (Completado) ---
    def _setup_mayor_tab(self):
        frame = ttk.Frame(self.frame_mayor)
        frame.pack(fill='both', expand=True, padx=10, pady=10)

        # Dropdown para seleccionar cuenta
        sel_frame = ttk.Frame(frame)
        sel_frame.pack(fill='x', pady=5)
        ttk.Label(sel_frame, text="Seleccione Cuenta:").pack(side='left', padx=5)
        
        self.mayor_cuenta_var = tk.StringVar()
        self.mayor_cuenta_combo = ttk.Combobox(sel_frame, textvariable=self.mayor_cuenta_var, state='readonly', width=50)
        self.mayor_cuenta_combo.pack(side='left', padx=5)
        self.mayor_cuenta_combo.bind("<<ComboboxSelected>>", lambda e: self.actualizar_mayor_cuenta_seleccionada())
        
        self.mayor_cuenta_map = {} # {Nombre: ID}
        self.mayor_saldo_label = ttk.Label(sel_frame, text="Saldo Actual: $0.00", font=('Inter', 10, 'bold'))
        self.mayor_saldo_label.pack(side='left', padx=20)
        
        # Treeview
        columns = ("Fecha", "Descripción", "Débito", "Crédito", "Saldo Acumulado")
        self.mayor_tree = ttk.Treeview(frame, columns=columns, show="headings")
        
        self.mayor_tree.heading("Fecha", text="Fecha")
        self.mayor_tree.heading("Descripción", text="Descripción")
        self.mayor_tree.heading("Débito", text="Débito ($)")
        self.mayor_tree.heading("Crédito", text="Crédito ($)")
        self.mayor_tree.heading("Saldo Acumulado", text="Saldo Acumulado ($)")
        
        self.mayor_tree.column("Fecha", width=120, anchor='center')
        self.mayor_tree.column("Descripción", width=300, anchor='w')
        self.mayor_tree.column("Débito", width=120, anchor='e')
        self.mayor_tree.column("Crédito", width=120, anchor='e')
        self.mayor_tree.column("Saldo Acumulado", width=150, anchor='e')
        
        self.mayor_tree.pack(side='left', fill='both', expand=True)
        
        # Scrollbar
        scrollbar = ttk.Scrollbar(frame, orient="vertical", command=self.mayor_tree.yview)
        self.mayor_tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side='right', fill='y')

    def actualizar_mayor_tab(self):
        cuentas = self.app.obtener_cuentas()
        nombres_cuentas = []
        self.mayor_cuenta_map = {}
        
        for id, nombre, tipo in cuentas:
            nombre_display = f"{nombre} ({tipo})"
            nombres_cuentas.append(nombre_display)
            self.mayor_cuenta_map[nombre_display] = id

        self.mayor_cuenta_combo['values'] = nombres_cuentas
        
        if not self.mayor_cuenta_var.get() and nombres_cuentas:
            self.mayor_cuenta_var.set(nombres_cuentas[0])
            self.actualizar_mayor_cuenta_seleccionada()
        elif self.mayor_cuenta_var.get():
            self.actualizar_mayor_cuenta_seleccionada()
        else:
            # Limpiar si no hay cuentas
            for item in self.mayor_tree.get_children(): self.mayor_tree.delete(item)
            self.mayor_saldo_label.config(text="Saldo Actual: $0.00")

    def actualizar_mayor_cuenta_seleccionada(self):
        cuenta_nombre = self.mayor_cuenta_var.get()
        cuenta_id = self.mayor_cuenta_map.get(cuenta_nombre)
        
        if not cuenta_id: return

        # Limpiar Treeview
        for item in self.mayor_tree.get_children():
            self.mayor_tree.delete(item)
            
        movimientos = self.app.obtener_movimientos_cuenta(cuenta_id)
        saldos_iniciales = self.app.calcular_saldos()
        
        # Determinar el tipo de saldo normal (Deudor o Acreedor)
        tipo_saldo = [data["tipo"] for id, data in saldos_iniciales.items() if id == cuenta_id][0]
        es_saldo_acreedor = tipo_saldo in self.app.tipos_saldo_acreedor

        saldo_acumulado = 0.0
        
        for movimiento in movimientos:
            fecha, desc, id_debito, id_credito, monto, ref_doc, ref_banco = movimiento
            
            monto_debito = monto if id_debito == cuenta_id else 0.0
            monto_credito = monto if id_credito == cuenta_id else 0.0
            
            # Cálculo del saldo
            if es_saldo_acreedor:
                # Saldo = Saldo anterior + Crédito - Débito
                saldo_acumulado += (monto_credito - monto_debito)
            else:
                # Saldo = Saldo anterior + Débito - Crédito
                saldo_acumulado += (monto_debito - monto_credito)

            self.mayor_tree.insert("", "end", values=(
                fecha.split()[0], desc, 
                f"{monto_debito:,.2f}", f"{monto_credito:,.2f}", f"{saldo_acumulado:,.2f}"
            ))
            
        # Actualizar etiqueta de Saldo Actual
        self.mayor_saldo_label.config(text=f"Saldo Actual: ${saldo_acumulado:,.2f}")
    
    # --- PESTAÑA: Balance de Comprobación (Completado) ---
    def _setup_balance_comp_tab(self):
        frame = ttk.Frame(self.frame_balance_comp)
        frame.pack(fill='both', expand=True, padx=10, pady=10)

        # Frame de botones
        btn_frame = ttk.Frame(frame)
        btn_frame.pack(fill='x', pady=5)
        ttk.Button(btn_frame, text="🔄 Recargar Balance", command=self.actualizar_balance_comp).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="📄 Exportar a PDF", command=self.exportar_balance_pdf_ui).pack(side='right', padx=5)
        ttk.Button(btn_frame, text="📊 Exportar a Excel", command=self.exportar_balance_excel_ui).pack(side='right', padx=5)

        # Treeview
        columns = ("Cuenta", "Tipo", "Debito", "Credito")
        self.balance_tree = ttk.Treeview(frame, columns=columns, show="headings")
        
        self.balance_tree.heading("Cuenta", text="Cuenta")
        self.balance_tree.heading("Tipo", text="Tipo")
        self.balance_tree.heading("Debito", text="Débito ($)")
        self.balance_tree.heading("Credito", text="Crédito ($)")
        
        self.balance_tree.column("Cuenta", width=300, anchor='w')
        self.balance_tree.column("Tipo", width=100, anchor='center')
        self.balance_tree.column("Debito", width=150, anchor='e')
        self.balance_tree.column("Credito", width=150, anchor='e')
        
        self.balance_tree.pack(side='top', fill='both', expand=True)
        
        # Totales
        self.balance_totales_label = ttk.Label(frame, text="Total Débito: $0.00 | Total Crédito: $0.00 | Cuadre: SI", font=('Inter', 12, 'bold'), anchor='e')
        self.balance_totales_label.pack(fill='x', pady=5)

    def actualizar_balance_comp(self):
        # Limpiar Treeview
        for item in self.balance_tree.get_children():
            self.balance_tree.delete(item)
            
        saldos = self.app.calcular_saldos()
        total_debitos_netos = 0.0
        total_creditos_netos = 0.0
        
        # Filtrar solo cuentas con saldo para mostrar en el Balance de Comprobación
        cuentas_con_saldo = {id: data for id, data in saldos.items() if abs(data["saldo"]) > 0.009}

        for id, data in cuentas_con_saldo.items():
            saldo = data["saldo"]
            debito_display = 0.00
            credito_display = 0.00
            
            # Si el saldo es positivo (Deudor), va a Débito
            if saldo > 0:
                debito_display = saldo
                total_debitos_netos += saldo
            # Si el saldo es negativo (Acreedor), va a Crédito (como valor absoluto)
            elif saldo < 0:
                credito_display = abs(saldo)
                total_creditos_netos += credito_display
            
            self.balance_tree.insert("", "end", values=(
                data["nombre"], data["tipo"], 
                f"{debito_display:,.2f}", f"{credito_display:,.2f}"
            ))

        # Actualizar etiqueta de totales
        cuadrado = "SI" if abs(total_debitos_netos - total_creditos_netos) < 0.01 else "NO"
        self.balance_totales_label.config(
            text=f"Total Débito: ${total_debitos_netos:,.2f} | Total Crédito: ${total_creditos_netos:,.2f} | Cuadre: {cuadrado}"
        )
        
    def exportar_balance_pdf_ui(self):
        ruta_archivo = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("Archivos PDF", "*.pdf")],
            title="Guardar Balance de Comprobación en PDF"
        )
        if ruta_archivo:
            exito, mensaje = self.app.exportar_balance_pdf(ruta_archivo)
            if exito:
                messagebox.showinfo("Éxito de Exportación", mensaje)
            else:
                messagebox.showerror("Error de Exportación", mensaje)

    def exportar_balance_excel_ui(self):
        ruta_archivo = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Archivos Excel", "*.xlsx")],
            title="Guardar Balance de Comprobación en Excel"
        )
        if ruta_archivo:
            exito, mensaje = self.app.exportar_balance_excel(ruta_archivo)
            if exito:
                messagebox.showinfo("Éxito de Exportación", mensaje)
            else:
                messagebox.showerror("Error de Exportación", mensaje)


    # --- PESTAÑA: Estado de Resultados (Completado) ---
    def _setup_estado_resultados_tab(self):
        frame = ttk.Frame(self.frame_estado_resultados)
        frame.pack(fill='both', expand=True, padx=10, pady=10)
        
        # Frame de botones
        btn_frame = ttk.Frame(frame)
        btn_frame.pack(fill='x', pady=5)
        ttk.Button(btn_frame, text="🔄 Recargar Reporte", command=self.actualizar_estado_resultados).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="📊 Exportar a Excel", command=self.exportar_estado_resultados_excel_ui).pack(side='right', padx=5)

        # Treeview (solo dos columnas: Concepto y Monto)
        columns = ("Concepto", "Monto")
        self.er_tree = ttk.Treeview(frame, columns=columns, show="headings")
        
        self.er_tree.heading("Concepto", text="Concepto")
        self.er_tree.heading("Monto", text="Monto ($)")
        
        self.er_tree.column("Concepto", width=400, anchor='w')
        self.er_tree.column("Monto", width=200, anchor='e')
        
        self.er_tree.pack(side='top', fill='both', expand=True)

    def actualizar_estado_resultados(self):
        # Limpiar Treeview
        for item in self.er_tree.get_children():
            self.er_tree.delete(item)
            
        datos = self.app.calcular_estado_resultados()
        
        # Separador para Ingresos
        self.er_tree.insert("", "end", values=("--- INGRESOS ---", ""), tags=('header',))
        
        # Insertar Ingresos
        self.er_tree.insert("", "end", values=("Ingresos Operacionales Netos", f"{datos['ingresos_totales']:,.2f}"))
        
        # Separador para Gastos
        self.er_tree.insert("", "end", values=("--- GASTOS ---", ""), tags=('header',))
        
        # Insertar Gastos
        saldos = self.app.calcular_saldos()
        gastos = {data["nombre"]: data["saldo"] for id, data in saldos.items() if data["tipo"] == "GASTO" and abs(data["saldo"]) > 0.009}
        
        for nombre, monto in gastos.items():
            self.er_tree.insert("", "end", values=(f"  - {nombre}", f"{monto:,.2f}"))

        # Fila de Totales de Gastos (para desglose visual)
        self.er_tree.insert("", "end", values=("Total Gastos", f"{datos['gastos_totales']:,.2f}"), tags=('subtotal',))
        
        # Fila de Utilidad/Pérdida
        concepto_final = "Utilidad Neta" if datos['utilidad_perdida_neta'] >= 0 else "Pérdida Neta"
        self.er_tree.insert("", "end", values=(concepto_final, f"{datos['utilidad_perdida_neta']:,.2f}"), tags=('final',))

        # Configurar estilos visuales
        self.er_tree.tag_configure('header', background='#D6EAF8', font=('Inter', 10, 'bold'))
        self.er_tree.tag_configure('subtotal', background='#EAECEE', font=('Inter', 10, 'bold'))
        self.er_tree.tag_configure('final', background='#A9CCE3', font=('Inter', 10, 'bold'))
        
    def exportar_estado_resultados_excel_ui(self):
        ruta_archivo = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Archivos Excel", "*.xlsx")],
            title="Guardar Estado de Resultados en Excel"
        )
        if ruta_archivo:
            exito, mensaje = self.app.exportar_estado_resultados_excel(ruta_archivo)
            if exito:
                messagebox.showinfo("Éxito de Exportación", mensaje)
            else:
                messagebox.showerror("Error de Exportación", mensaje)


    # --- PESTAÑA: Balance General (Completado) ---
    def _setup_balance_general_tab(self):
        frame = ttk.Frame(self.frame_balance_general)
        frame.pack(fill='both', expand=True, padx=10, pady=10)
        
        # Frame de botones
        btn_frame = ttk.Frame(frame)
        btn_frame.pack(fill='x', pady=5)
        ttk.Button(btn_frame, text="🔄 Recargar Reporte", command=self.actualizar_balance_general).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="📊 Exportar a Excel", command=self.exportar_balance_general_excel_ui).pack(side='right', padx=5)

        # Treeview (solo dos columnas: Concepto y Monto)
        columns = ("Concepto", "Monto")
        self.bg_tree = ttk.Treeview(frame, columns=columns, show="headings")
        
        self.bg_tree.heading("Concepto", text="Concepto")
        self.bg_tree.heading("Monto", text="Monto ($)")
        
        self.bg_tree.column("Concepto", width=400, anchor='w')
        self.bg_tree.column("Monto", width=200, anchor='e')
        
        self.bg_tree.pack(side='top', fill='both', expand=True)
        
        # Etiqueta de la Ecuación Contable
        self.ecuacion_label = ttk.Label(frame, text="Ecuación Contable: Activo ($0.00) = Pasivo + Capital ($0.00)", font=('Inter', 10, 'bold'), anchor='e')
        self.ecuacion_label.pack(fill='x', pady=5)

    def actualizar_balance_general(self):
        # Limpiar Treeview
        for item in self.bg_tree.get_children():
            self.bg_tree.delete(item)
            
        datos = self.app.calcular_balance_general()
        saldos = self.app.calcular_saldos()
        
        # --- ACTIVO ---
        self.bg_tree.insert("", "end", values=("--- ACTIVO (Recursos) ---", ""), tags=('header',))
        activos = {data["nombre"]: data["saldo"] for id, data in saldos.items() if data["tipo"] == "ACTIVO" and abs(data["saldo"]) > 0.009}
        
        for nombre, monto in activos.items():
            self.bg_tree.insert("", "end", values=(f"  - {nombre}", f"{monto:,.2f}"))
            
        self.bg_tree.insert("", "end", values=("TOTAL ACTIVO", f"{datos['activo']:,.2f}"), tags=('total',))
        self.bg_tree.insert("", "end", values=("", ""))

        # --- PASIVO Y CAPITAL ---
        self.bg_tree.insert("", "end", values=("--- PASIVO Y CAPITAL (Fuentes de Financiación) ---", ""), tags=('header',))
        
        # PASIVO
        self.bg_tree.insert("", "end", values=("PASIVO (Deudas):", ""), tags=('subheader',))
        pasivos = {data["nombre"]: abs(data["saldo"]) for id, data in saldos.items() if data["tipo"] == "PASIVO" and abs(data["saldo"]) > 0.009}
        
        for nombre, monto in pasivos.items():
            self.bg_tree.insert("", "end", values=(f"  - {nombre}", f"{monto:,.2f}"))
            
        self.bg_tree.insert("", "end", values=("Total Pasivo", f"{datos['pasivo']:,.2f}"), tags=('subtotal',))
        self.bg_tree.insert("", "end", values=("", ""))

        # CAPITAL (PATRIMONIO)
        self.bg_tree.insert("", "end", values=("CAPITAL (Patrimonio):", ""), tags=('subheader',))
        capitales = {data["nombre"]: abs(data["saldo"]) for id, data in saldos.items() if data["tipo"] == "CAPITAL" and abs(data["saldo"]) > 0.009}
        
        for nombre, monto in capitales.items():
            self.bg_tree.insert("", "end", values=(f"  - {nombre}", f"{monto:,.2f}"))
            
        # Resultado del Ejercicio
        concepto_resultado = "Utilidad del Ejercicio" if datos['utilidad_neta'] >= 0 else "Pérdida del Ejercicio"
        self.bg_tree.insert("", "end", values=(concepto_resultado, f"{datos['utilidad_neta']:,.2f}"))
        
        self.bg_tree.insert("", "end", values=("TOTAL PASIVO + CAPITAL (Ajustado)", f"{datos['pasivo_mas_capital_ajustado']:,.2f}"), tags=('final',))
        
        # Configurar estilos visuales
        self.bg_tree.tag_configure('header', background='#D6EAF8', font=('Inter', 10, 'bold'))
        self.bg_tree.tag_configure('subheader', font=('Inter', 10, 'bold'))
        self.bg_tree.tag_configure('subtotal', background='#EAECEE', font=('Inter', 10, 'bold'))
        self.bg_tree.tag_configure('total', background='#A9CCE3', font=('Inter', 10, 'bold'))
        self.bg_tree.tag_configure('final', background='#A9CCE3', font=('Inter', 10, 'bold'))

        # Actualizar etiqueta de la Ecuación Contable
        cuadrado = "SI" if abs(datos['activo'] - datos['pasivo_mas_capital_ajustado']) < 0.01 else "NO"
        self.ecuacion_label.config(
            text=f"Ecuación Contable: Activo (${datos['activo']:,.2f}) = Pasivo + Capital (${datos['pasivo_mas_capital_ajustado']:,.2f}) | Cuadre: {cuadrado}"
        )
        
    def exportar_balance_general_excel_ui(self):
        ruta_archivo = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Archivos Excel", "*.xlsx")],
            title="Guardar Balance General en Excel"
        )
        if ruta_archivo:
            exito, mensaje = self.app.exportar_balance_general_excel(ruta_archivo)
            if exito:
                messagebox.showinfo("Éxito de Exportación", mensaje)
            else:
                messagebox.showerror("Error de Exportación", mensaje)

    # --- PESTAÑA: Importar Datos (Completado) ---
    def _setup_importar_tab(self):
        frame = ttk.Frame(self.frame_importar)
        frame.pack(fill='both', expand=True, padx=10, pady=10)
        
        ttk.Label(frame, text="📥 Importar Transacciones desde Excel", font=('Inter', 16, 'bold')).pack(pady=10)
        
        self.ruta_archivo_var = tk.StringVar()
        
        # Frame de selección de archivo
        file_frame = ttk.LabelFrame(frame, text="Seleccionar Archivo (.xlsx)", padding=10)
        file_frame.pack(fill='x', padx=10, pady=5)
        
        ttk.Entry(file_frame, textvariable=self.ruta_archivo_var, width=80).pack(side='left', padx=5, fill='x', expand=True)
        ttk.Button(file_frame, text="Buscar Archivo...", command=self.seleccionar_archivo_excel).pack(side='left', padx=5)

        # Botón de importación
        ttk.Button(frame, text="Ejecutar Importación de Transacciones", command=self.ejecutar_importacion_excel).pack(pady=20)
        
        # Instrucciones
        instrucciones = (
            "⚠️ Formato Requerido del Archivo Excel (Primera Fila/Encabezados):\n"
            "1. DESCRIPCION (Texto)\n"
            "2. CUENTA_DEBITO (Nombre exacto de la cuenta existente)\n"
            "3. CUENTA_CREDITO (Nombre exacto de la cuenta existente)\n"
            "4. MONTO (Número positivo)\n"
            "5. REF_DOC (Opcional, Texto)\n"
            "6. REF_BANCO (Opcional, Texto)"
        )
        ttk.Label(frame, text=instrucciones, justify=tk.LEFT, background='#ecf0f1', padding=10).pack(fill='x', padx=10, pady=10)

    def seleccionar_archivo_excel(self):
        ruta = filedialog.askopenfilename(
            defaultextension=".xlsx",
            filetypes=[("Archivos Excel", "*.xlsx"), ("Archivos Excel (Legacy)", "*.xls")],
            title="Seleccionar Archivo de Transacciones"
        )
        if ruta:
            self.ruta_archivo_var.set(ruta)

    def ejecutar_importacion_excel(self):
        ruta = self.ruta_archivo_var.get()
        if not ruta:
            messagebox.showwarning("Advertencia", "Debe seleccionar un archivo Excel para importar.")
            return

        exito, mensaje = self.app.importar_transacciones_excel(ruta)
        
        if exito:
            messagebox.showinfo("Importación Exitosa", mensaje)
            self.recargar_vistas_contables()
            self.actualizar_combos_transaccion()
        else:
            messagebox.showerror("Error de Importación", mensaje)


# --- 5. EJECUCIÓN PRINCIPAL ---

if __name__ == '__main__':
    # Usar el try-except para asegurar que la conexión se cierre incluso si hay un error
    app_backend = ContabilidadApp()
    try:
        root = tk.Tk()
        app_gui = ContabilidadGUI(root, app_backend)
        root.mainloop()
    except Exception as e:
        print(f"Error en la ejecución de la GUI: {e}")
        messagebox.showerror("Error Crítico de la Aplicación", f"Ocurrió un error inesperado: {e}")
    finally:
        app_backend.cerrar_conexion()
        print("Conexión a la base de datos cerrada.")